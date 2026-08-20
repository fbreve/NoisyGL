import os
import re
import csv
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX_PATH = r'results/NoisyGL.xlsx'
CSV_PATH = r'log/lnpcc_results_20260421_172348.csv'
INSTANCE_CSV_PATH = r'log/instance_final_results.csv'
TARGET_XLSX = r'results/results.xlsx'

all_results = {}
def add_result(method, dataset, ntype, rate, data):
    if data is None: return
    if method not in all_results: all_results[method] = {}
    if dataset not in all_results[method]: all_results[method][dataset] = {}
    if ntype not in all_results[method][dataset]: all_results[method][dataset][ntype] = {}
    all_results[method][dataset][ntype][rate] = data

def parse_noise_label_excel(noise_raw):
    noise_raw = str(noise_raw).strip().lower()
    if 'clean' in noise_raw: return 'clean', 0.0
    m = re.search(r'(\d+)\s*%?\s+([a-z-]+)', noise_raw)
    if m:
        rate = float(m.group(1)) / 100.0
        ntype = m.group(2)
        if 'asym' in ntype: ntype = 'random'
        return ntype, rate
    return None, None

def parse_val_with_std(val_str):
    if not val_str or str(val_str).lower() == 'nan': return None
    nums = re.findall(r'[\d.]+', str(val_str))
    if len(nums) >= 2: return {'mean': float(nums[0]), 'std': float(nums[1])}
    elif len(nums) == 1: return {'mean': float(nums[0]), 'std': 0.0}
    return None

EXCEL_DS_MAP = {
    'cora': 'cora', 'citeseer': 'citeseer', 'pubmed': 'pubmed',
    'amazon-c': 'amazoncom', 'amazon-p': 'amazonpho', 'dblp': 'dblp',
    'blogcatalog': 'blogcatalog', 'flickr': 'flickr', 'amz-rat.': 'amazon-ratings',
    'roman-emp.': 'roman-empire', 'a-computers': 'amazoncom', 'a-photo': 'amazonpho',
    'a-photos': 'amazonpho', 'a-ratings': 'amazon-ratings', 'empire': 'roman-empire',
    'amazonpho': 'amazonpho', 'amazoncom': 'amazoncom',
}

wb_raw = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
sheet = wb_raw.active
rows = list(sheet.iter_rows(values_only=True))
header_row_idx = -1
for i in range(20):
    if len(rows[i]) > 2 and str(rows[i][2]).strip() == 'Dataset':
        header_row_idx = i
        break

methods_found = {}
for col_idx in range(4, len(rows[header_row_idx])):
    m_name = str(rows[header_row_idx][col_idx]).strip()
    if m_name and m_name != 'None' and m_name != 'nan': 
        methods_found[col_idx] = m_name

blocks = []
current_block = None
for idx in range(header_row_idx + 1, len(rows)):
    row = rows[idx]
    if len(row) < 4: continue
    noise_raw = str(row[3]).strip().lower() if row[3] is not None else ''
    if 'clean' in noise_raw:
        current_block = {'dataset': None, 'rows': []}
        blocks.append(current_block)
    if current_block is not None:
        current_block['rows'].append(idx)
        ds_raw = str(row[2]).strip() if row[2] is not None else ''
        if ds_raw and ds_raw != 'None' and ds_raw != 'Dataset':
            ds_candidate = ds_raw.split()[-1].lower()
            current_block['dataset'] = EXCEL_DS_MAP.get(ds_candidate, ds_candidate)

for b in blocks:
    dataset = b['dataset']
    if not dataset: continue
    for idx in b['rows']:
        row = rows[idx]
        ntype, rate = parse_noise_label_excel(str(row[3]))
        if ntype is not None:
            for col_idx, m_name in methods_found.items():
                data = parse_val_with_std(str(row[col_idx]))
                add_result(m_name, dataset, ntype, rate, data)

with open(CSV_PATH, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        dataset = row['dataset'].strip().lower()
        label = row['noise_rate'].strip().lower()
        if label.startswith('clean'): ntype, rate = 'clean', 0.0
        else:
            parts = label.split('_')
            ntype, rate = parts[0], float(parts[1])
        data = {'mean': float(row['lnpcc_mean']) * 100.0, 'std': float(row['lnpcc_std']) * 100.0}
        add_result('PCC+GCN', dataset, ntype, rate, data)

with open(INSTANCE_CSV_PATH, newline='', encoding='utf-8') as f:
    for row in csv.DictReader(f):
        dataset = row['dataset'].strip().lower()
        rate = float(row['noise_rate'])
        method_raw = row['method'].strip().lower()
        if method_raw == 'lnpcc': method = 'PCC+GCN'
        elif method_raw == 'gcn': method = 'GCN'
        elif method_raw == 'cp': method = 'CP'
        elif method_raw == 'nrgnn': method = 'NRGNN'
        elif method_raw == 'pignn': method = 'PIGNN'
        else: method = method_raw.upper()
        data = {'mean': float(row['acc_mean']) * 100.0, 'std': float(row['acc_std']) * 100.0, 'cpu': float(row['cpu_s_avg']), 'gpu': float(row['gpu_s_avg'])}
        add_result(method, dataset, 'instance', rate, data)

DATASET_ORDER = ['cora', 'citeseer', 'pubmed', 'blogcatalog', 'flickr', 'dblp', 'amazoncom', 'amazonpho', 'amazon-ratings', 'roman-empire']
DATASET_LABELS = {
    'cora': 'Cora', 'citeseer': 'CiteSeer', 'pubmed': 'PubMed',
    'amazoncom': 'Amazon-C', 'amazonpho': 'Amazon-P', 'dblp': 'DBLP',
    'blogcatalog': 'BlogCatalog', 'flickr': 'Flickr',
    'amazon-ratings': 'Amz-Rat.', 'roman-empire': 'Roman-Emp.'
}
instance_methods = ['GCN', 'CP', 'NRGNN', 'PIGNN', 'PCC+GCN']

# Styles
hdr_fill = PatternFill("solid", fgColor="1F3864")
subhdr_fill = PatternFill("solid", fgColor="2F5597")
hdr_font = Font(bold=True, color="FFFFFF", size=10)
best_fill = PatternFill("solid", fgColor="A8D08D")
gain_pos_fill = PatternFill("solid", fgColor="E2EFDA")
gain_neg_fill = PatternFill("solid", fgColor="FFC7CE")
bold_font = Font(bold=True, size=10)
thin = Side(style='thin', color='BFBFBF')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

# Load target workbook
if os.path.exists(TARGET_XLSX):
    wb = openpyxl.load_workbook(TARGET_XLSX)
else:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

# ---------------------------------------------------------------------------
# 1. Sheet: Instance Accuracy
# ---------------------------------------------------------------------------
sheet_name = 'Instance Accuracy'
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)

ws.append(['Mean Accuracy Comparison on Instance Noise (Rates 0.1 - 0.5 Average)'])
ws.cell(1, 1).font = Font(bold=True, size=12)
header = ['Method'] + [DATASET_LABELS[d] for d in DATASET_ORDER] + ['Average']
ws.append(header)
for col_idx in range(1, len(header) + 1):
    c = ws.cell(2, col_idx)
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

data_rows = []
for m in instance_methods:
    row_vals = [m]
    ds_means = []
    for ds in DATASET_ORDER:
        vals = [all_results[m][ds]['instance'][r]['mean'] for r in [0.1, 0.2, 0.3, 0.4, 0.5] if m in all_results and ds in all_results[m] and 'instance' in all_results[m][ds] and r in all_results[m][ds]['instance']]
        if vals:
            v = np.mean(vals)
            ds_means.append(v)
            row_vals.append(round(v, 2))
        else:
            row_vals.append(None)
    row_vals.append(round(np.mean(ds_means), 2) if ds_means else None)
    data_rows.append(row_vals)
    ws.append(row_vals)

# Formatting
for r in range(3, 3 + len(data_rows)):
    for col_idx in range(1, len(header) + 1):
        c = ws.cell(r, col_idx)
        c.border = border
        c.alignment = center if col_idx > 1 else left_align
        if col_idx > 1:
            col_vals = [data_rows[i][col_idx-1] for i in range(len(data_rows)) if data_rows[i][col_idx-1] is not None]
            if col_vals and c.value is not None and abs(c.value - max(col_vals)) < 1e-5:
                c.fill = best_fill
                c.font = bold_font

ws.column_dimensions['A'].width = 15
for i in range(len(DATASET_ORDER) + 1):
    ws.column_dimensions[get_column_letter(i+2)].width = 12
ws.freeze_panes = 'B3'

# ---------------------------------------------------------------------------
# 2. Sheet: Instance Training Time
# ---------------------------------------------------------------------------
sheet_name = 'Instance Training Time'
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)

ws.append(['Training Time on Instance Noise (Stacked CPU, GPU, Total in seconds)'])
ws.cell(1, 1).font = Font(bold=True, size=12)
header = ['Method', 'Metric'] + [DATASET_LABELS[d] for d in DATASET_ORDER] + ['Average']
ws.append(header)
for col_idx in range(1, len(header) + 1):
    c = ws.cell(2, col_idx)
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

for m in instance_methods:
    for metric, label in [('cpu', 'CPU'), ('gpu', 'GPU'), ('total', 'Total')]:
        row_vals = [m if label == 'CPU' else '', label]
        ds_times = []
        for ds in DATASET_ORDER:
            times = []
            for r in [0.1, 0.2, 0.3, 0.4, 0.5]:
                if m in all_results and ds in all_results[m] and 'instance' in all_results[m][ds] and r in all_results[m][ds]['instance']:
                    e = all_results[m][ds]['instance'][r]
                    if metric == 'cpu': times.append(e.get('cpu', 0))
                    elif metric == 'gpu': times.append(e.get('gpu', 0))
                    elif metric == 'total': times.append(e.get('cpu', 0) + e.get('gpu', 0))
            if times:
                v = np.mean(times)
                ds_times.append(v)
                row_vals.append(round(v, 2))
            else:
                row_vals.append(None)
        row_vals.append(round(np.mean(ds_times), 2) if ds_times else None)
        ws.append(row_vals)
        r = ws.max_row
        for col_idx in range(1, len(header) + 1):
            c = ws.cell(r, col_idx)
            c.border = border
            c.alignment = center if col_idx > 2 else left_align
            if label == 'Total':
                c.font = bold_font
                if col_idx == 2:
                    c.fill = PatternFill("solid", fgColor="F2F2F2")

ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 10
for i in range(len(DATASET_ORDER) + 1):
    ws.column_dimensions[get_column_letter(i+3)].width = 12
ws.freeze_panes = 'C3'

# ---------------------------------------------------------------------------
# 3. Sheet: Summary by Noise Rate (Acc)
# ---------------------------------------------------------------------------
sheet_name = 'Summary by Noise Rate (Acc)'
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)

ws.append(['Global Mean Accuracy across 10 Datasets by Noise Rate'])
ws.cell(1, 1).font = Font(bold=True, size=12)

rates = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5]
header = ['Noise Type', 'Method', 'Clean (0.0)', '0.1', '0.2', '0.3', '0.4', '0.5', 'Avg (0.1-0.5)', 'Overall Avg']
ws.append(header)
for col_idx in range(1, len(header) + 1):
    c = ws.cell(2, col_idx)
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

for nt in ['instance', 'uniform', 'pair', 'random']:
    methods_to_show = instance_methods if nt == 'instance' else sorted(all_results.keys())
    start_r = ws.max_row + 1
    table_rows = []
    for m in methods_to_show:
        row_vals = [nt.upper(), m]
        r_means = {}
        for r in rates:
            curr_nt, curr_r = ('clean', 0.0) if r == 0.0 else (nt, r)
            ds_vals = [all_results[m][d][curr_nt][curr_r]['mean'] for d in DATASET_ORDER if m in all_results and d in all_results[m] and curr_nt in all_results[m][d] and curr_r in all_results[m][d][curr_nt]]
            if ds_vals:
                r_means[r] = np.mean(ds_vals)
                row_vals.append(round(np.mean(ds_vals), 2))
            else:
                row_vals.append(None)
        noisy_avg = [r_means[r] for r in [0.1, 0.2, 0.3, 0.4, 0.5] if r in r_means]
        all_avg = [r_means[r] for r in rates if r in r_means]
        row_vals.append(round(np.mean(noisy_avg), 2) if noisy_avg else None)
        row_vals.append(round(np.mean(all_avg), 2) if all_avg else None)
        table_rows.append(row_vals)
        ws.append(row_vals)
    
    end_r = ws.max_row
    # Highlight best in this noise group
    for r in range(start_r, end_r + 1):
        for col_idx in range(1, len(header) + 1):
            c = ws.cell(r, col_idx)
            c.border = border
            c.alignment = center if col_idx > 2 else left_align
            if col_idx >= 3:
                vals = [table_rows[i][col_idx-1] for i in range(len(table_rows)) if table_rows[i][col_idx-1] is not None]
                if vals and c.value is not None and abs(c.value - max(vals)) < 1e-5:
                    c.fill = best_fill
                    c.font = bold_font

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 15
for i in range(len(header) - 2):
    ws.column_dimensions[get_column_letter(i+3)].width = 14
ws.freeze_panes = 'C3'

# ---------------------------------------------------------------------------
# 4. Sheet: Summary by Noise Rate (Gain)
# ---------------------------------------------------------------------------
sheet_name = 'Summary by Noise Rate (Gain)'
if sheet_name in wb.sheetnames: del wb[sheet_name]
ws = wb.create_sheet(title=sheet_name)

ws.append(['Global Mean Accuracy Delta vs GCN across 10 Datasets (percentage points)'])
ws.cell(1, 1).font = Font(bold=True, size=12)

header = ['Noise Type', 'Method', 'Clean (0.0)', '0.1', '0.2', '0.3', '0.4', '0.5', 'Average Delta']
ws.append(header)
for col_idx in range(1, len(header) + 1):
    c = ws.cell(2, col_idx)
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = center; c.border = border

for nt in ['instance', 'uniform', 'pair', 'random']:
    methods_to_show = instance_methods if nt == 'instance' else sorted(all_results.keys())
    for m in methods_to_show:
        row_vals = [nt.upper(), m]
        deltas = {}
        for r in rates:
            curr_nt, curr_r = ('clean', 0.0) if r == 0.0 else (nt, r)
            ds_deltas = []
            for d in DATASET_ORDER:
                if m in all_results and d in all_results[m] and curr_nt in all_results[m][d] and curr_r in all_results[m][d][curr_nt]:
                    if 'GCN' in all_results and d in all_results['GCN'] and curr_nt in all_results['GCN'][d] and curr_r in all_results['GCN'][d][curr_nt]:
                        delta = all_results[m][d][curr_nt][curr_r]['mean'] - all_results['GCN'][d][curr_nt][curr_r]['mean']
                        ds_deltas.append(delta)
            if ds_deltas:
                v = np.mean(ds_deltas)
                deltas[r] = v
                row_vals.append(round(v, 2))
            else:
                row_vals.append(None)
        all_d = [deltas[r] for r in rates if r in deltas]
        row_vals.append(round(np.mean(all_d), 2) if all_d else None)
        ws.append(row_vals)
        r_idx = ws.max_row
        for col_idx in range(1, len(header) + 1):
            c = ws.cell(r_idx, col_idx)
            c.border = border
            c.alignment = center if col_idx > 2 else left_align
            if col_idx >= 3 and c.value is not None:
                if c.value > 0.5:
                    c.fill = gain_pos_fill
                    if m == 'PCC+GCN': c.font = bold_font
                elif c.value < -0.5:
                    c.fill = gain_neg_fill

ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 15
for i in range(len(header) - 2):
    ws.column_dimensions[get_column_letter(i+3)].width = 14
ws.freeze_panes = 'C3'

# Save updated workbook
wb.save(TARGET_XLSX)
print(f"Successfully added summary sheets to {TARGET_XLSX}")
print("Sheets now in workbook:", wb.sheetnames)
