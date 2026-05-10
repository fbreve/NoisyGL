"""
generate_tables.py
==================
Gera tabelas de resultados do LN-PCC em três formatos simultâneos:
  - Terminal (texto formatado)
  - Excel (.xlsx, com formatação)
  - LaTeX (.tex, estilo paper)
Tabelas geradas:
  1. Resultados absolutos   — acurácia (%) com ± std
  2. Ganho relativo ao GCN  — Δ% em relação ao GCN
  3. Hiperparâmetros        — parâmetros ajustados por dataset
Estrutura das tabelas de resultados:
  Linhas  : (noise_type, noise_rate)  +  linha "Média" ao final
  Colunas : um dataset por coluna
Uso básico (sem argumentos — busca automaticamente o CSV mais recente):
  python results/generate_tables.py
Uso avançado:
  python results/generate_tables.py --csv log/meu_arquivo.csv
  python results/generate_tables.py --methods lnpcc gcn
  python results/generate_tables.py --noise_types uniform pair
"""
import argparse
import glob
import os
import re
import sys
import csv
import math
from collections import defaultdict
from pathlib import Path
# ---------------------------------------------------------------------------
# Depend on pandas / openpyxl (soft requirement for Excel)
# ---------------------------------------------------------------------------
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    pd = None
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False
# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
SCRIPT_DIR  = Path(__file__).parent          # .../NoisyGL/
PROJECT_DIR = SCRIPT_DIR                     # .../NoisyGL/
LOG_DIR     = PROJECT_DIR / "log"
CONFIG_DIR  = PROJECT_DIR / "config" / "lnpcc"
HPO_DB_PATH = LOG_DIR / "hpo_db.json"
DATASET_ORDER = [
    'cora', 'citeseer', 'pubmed',
    'amazoncom', 'amazonpho', 'dblp',
    'blogcatalog', 'flickr',
    'amazon-ratings', 'roman-empire',
]
DATASET_LABELS = {
    'cora':          'Cora',
    'citeseer':      'CiteSeer',
    'pubmed':        'PubMed',
    'amazoncom':     'Amazon-C',
    'amazonpho':     'Amazon-P',
    'dblp':          'DBLP',
    'blogcatalog':   'BlogCatalog',
    'flickr':        'Flickr',
    'amazon-ratings':'Amz-Rat.',
    'roman-empire':  'Roman-Emp.',
}
NOISE_TYPE_ORDER = ['clean', 'uniform', 'pair', 'random']
NOISE_TYPE_LABELS = {
    'clean':    'Clean',
    'uniform':  'Uniform',
    'pair':     'Pair',
    'random':   'Random',
}
# Hyperparameters present in hpo_db.json and their display labels
HYPERPARAM_DISPLAY = {
    'knn_mode':     'KNN Mode',
    'k':            'k',
    'dexp':         'd_exp',
    'p_grd':        'p_grd',
    'unc_rem':      'Rem Unc',
    'unc_rel':      'Rel Unc',
    'dropout':      'Dropout',
    'n_hidden':     'Hidden',
    'n_layer':      'Layers',
    'lr':           'LR',
    'weight_decay': 'WD',
}
# Hyperparameters to format as integers
HYPERPARAM_INT = {'k', 'n_hidden', 'n_layer'}
# Hyperparameters to format with 4 sig-figs
HYPERPARAM_FLOAT4 = {'p_grd', 'unc_rem', 'unc_rel', 'dropout', 'lr', 'weight_decay', 'dexp'}
METHOD_LABELS = {
    'lnpcc': 'LN-PCC',
    'gcn':   'GCN',
    'nrgnn': 'NRGNN',
    'rtgnn': 'RTGNN',
    'cp':    'CP',
    'clnode':'CLNode',
    'pignn': 'PIGNN',
    'dgnn':  'DGNN',
    'rncgln':'RNCGLN',
    'unionnet':'UnionNET',
    'cgnn':  'CGNN',
    'cr-gnn': 'CR-GNN',
}
# ---------------------------------------------------------------------------
# Helper: find latest CSV
# ---------------------------------------------------------------------------
def find_latest_csv(log_dir: Path) -> Path:
    """Return the most recently modified lnpcc_results_*.csv in log_dir."""
    pattern = str(log_dir / "lnpcc_results_*.csv")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(
            f"No lnpcc_results_*.csv found in {log_dir}.\n"
            "Run a benchmark first or pass --csv explicitly."
        )
    return Path(max(files, key=os.path.getmtime))
# ---------------------------------------------------------------------------
# CSV parsing
# ---------------------------------------------------------------------------
def parse_noise_label(label: str):
    """'uniform_0.3' → ('uniform', 0.3);  'clean_0.0' → ('clean', 0.0)."""
    label = label.strip().lower()
    if label.startswith('clean'):
        return 'clean', 0.0
    m = re.match(r'^([a-z]+)_([0-9.]+)$', label)
    if m:
        return m.group(1), float(m.group(2))
    raise ValueError(f"Cannot parse noise label: {label!r}")
def load_csv(path: Path) -> dict:
    """
    Load a lnpcc_results_*.csv.
    Returns
    -------
    dict: {(method, dataset, noise_type, noise_rate): (mean%, std%)}
    """
    data = {}
    with open(path, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            dataset = row['dataset'].strip().lower()
            try:
                noise_type, noise_rate = parse_noise_label(row['noise_rate'])
            except ValueError:
                continue
            for col, val in row.items():
                if not col.endswith('_mean') or not val.strip():
                    continue
                base = col[:-5]  # strip '_mean'
                # Resolve method name
                if base.startswith('lnpcc_'):
                    method = base  # keep prefixed for multi-variant (legacy)
                else:
                    method = base  # 'gcn', 'lnpcc', etc.
                std_col = f'{base}_std'
                try:
                    mean_v = float(val) * 100.0
                    std_v  = float(row.get(std_col, 0)) * 100.0
                except (ValueError, TypeError):
                    continue
                # Normalise 'lnpcc' (un-prefixed)
                key = (method, dataset, noise_type, noise_rate)
                data[key] = (mean_v, std_v)
    return data

def load_external_baseline_xlsx(path: Path) -> dict:
    """
    Load ALL baseline results from NoisyGL.xlsx.
    Returns
    -------
    dict: {(method_id, dataset, noise_type, noise_rate): (mean%, std%)}
    """
    if not HAS_PANDAS:
        print("[AVISO] pandas não instalado — não foi possível carregar o baseline Excel.")
        return {}
    if not path.exists():
        print(f"[AVISO] Baseline Excel não encontrada: {path}")
        return {}

    data = {}
    try:
        df = pd.read_excel(path, sheet_name=0, header=None) # Sheet 0 is usually Planilha1
    except Exception as e:
        print(f"[ERRO] Falha ao ler Excel: {e}")
        return {}

    # Find header row
    header_row_idx = -1
    for i in range(15):
        if str(df.iloc[i, 2]).strip() == "Dataset":
            header_row_idx = i
            break
    if header_row_idx == -1: return {}

    # Identify method columns
    methods_found = {}
    for col_idx in range(4, len(df.columns)):
        m_name = str(df.iloc[header_row_idx, col_idx]).strip().lower()
        if m_name and m_name != "nan":
             # Normalize name to match our keys (e.g. 'cr-gnn' or 'clnode')
             m_id = m_name.replace(' ', '-').replace('_', '-')
             methods_found[col_idx] = m_id

    ds_map = {
        'a-computers': 'amazoncom', 'a-photo': 'amazonpho', 'a-photos': 'amazonpho',
        'a-ratings': 'amazon-ratings', 'empire': 'roman-empire', 'cora': 'cora',
        'citeseer': 'citeseer', 'pubmed': 'pubmed', 'dblp': 'dblp',
        'blogcatalog': 'blogcatalog', 'flickr': 'flickr',
    }

    # Block parsing logic
    blocks = []
    current_block = None
    for idx in range(header_row_idx + 1, len(df)):
        row = df.iloc[idx]
        noise_raw = str(row[3]).strip().lower() if pd.notna(row[3]) else ""
        if 'clean' in noise_raw:
            current_block = {'dataset': None, 'rows': []}
            blocks.append(current_block)
        if current_block is not None:
            current_block['rows'].append(idx)
            ds_raw = str(row[2]).strip() if pd.notna(row[2]) else ""
            if ds_raw and ds_raw != "nan" and ds_raw != "Dataset":
                ds_parts = ds_raw.split()
                ds_candidate = ds_parts[-1].lower() if ds_parts else ""
                if ds_candidate in ds_map: current_block['dataset'] = ds_map[ds_candidate]
                elif any(ds_candidate == d.lower() for d in DATASET_ORDER): current_block['dataset'] = ds_candidate

    for block in blocks:
        dataset = block['dataset']
        if not dataset: continue
        for idx in block['rows']:
            row = df.iloc[idx]
            noise_raw = str(row[3]).strip().lower()
            if 'clean' in noise_raw: nt, nr = 'clean', 0.0
            else:
                m = re.search(r'(\d+)\s*%?\s+([a-z-]+)', noise_raw)
                if m:
                    nr, nt = float(m.group(1)) / 100.0, m.group(2)
                    if nt in ['asymmetric', 'pair-asym']: nt = 'random'
                else: continue
            
            for col_idx, m_id in methods_found.items():
                raw_str = str(row[col_idx]).strip()
                if not raw_str or raw_str == "nan": continue
                nums = re.findall(r'[\d.]+', raw_str)
                if nums:
                    mean_v = float(nums[0])
                    std_v = float(nums[1]) if len(nums) > 1 else 0.0
                    data[(m_id, dataset, nt, nr)] = (mean_v, std_v)
    return data
# ---------------------------------------------------------------------------
# Build pivot table
# ---------------------------------------------------------------------------
def build_pivot(data: dict, method: str, datasets: list, noise_types: list, noise_rates: list):
    """
    Returns a list of rows for a single method benchmark.
    """
    rows = []
    for nt in noise_types:
        rates = [0.0] if nt == 'clean' else [nr for nr in noise_rates if nr > 0.0]
        for nr in rates:
            row = {'noise_type': nt, 'noise_rate': nr}
            for ds in datasets:
                row[ds] = data.get((method, ds, nt, nr))
            rows.append(row)
    # ── Mean row (standard average across all scenarios) ──
    mean_row = {'noise_type': 'MEAN', 'noise_rate': None}
    for ds in datasets:
        vals = [r[ds][0] for r in rows if r[ds] is not None]
        if vals:
            mean_row[ds] = (sum(vals) / len(vals), None)
        else:
            mean_row[ds] = None
    rows.append(mean_row)
    return rows

def build_summary_table(data: dict, methods: list, datasets: list):
    """
    Table 1: Each row is a method, each column is a dataset.
    Value is the mean accuracy across all scenarios (clean + all noise).
    Ordering: GCN first, others alphabetical, LN-PCC last.
    """
    # Define order
    others = sorted([m for m in methods if m not in ['gcn', 'lnpcc']])
    ordered_methods = []
    if 'gcn' in methods: ordered_methods.append('gcn')
    ordered_methods.extend(others)
    if 'lnpcc' in methods: ordered_methods.append('lnpcc')

    summary_rows = []
    for m in ordered_methods:
        row = {'method': m}
        dataset_means = []
        for ds in datasets:
            # Find all scenarios for this (method, dataset)
            scenarios = [data[k][0] for k in data if k[0] == m and k[1] == ds]
            if scenarios:
                m_val = sum(scenarios) / len(scenarios)
                row[ds] = m_val
                dataset_means.append(m_val)
            else:
                row[ds] = None
        # Global mean for this method across datasets
        row['GLOBAL_MEAN'] = sum(dataset_means) / len(dataset_means) if dataset_means else None
        summary_rows.append(row)
    return summary_rows

def build_ranking_table(summary_rows: list, datasets: list):
    """
    Table 2: Ranking of Table 1.
    The final column shows the Mean Rank (average of ranks across datasets).
    """
    ranking_rows = []
    for row in summary_rows:
        ranking_rows.append({'method': row['method'], 'ranks': {}})
    
    # 1. Rank datasets individually
    for ds in datasets:
        # Get all non-None values for this column
        vals = []
        for i, row in enumerate(summary_rows):
            if row.get(ds) is not None:
                vals.append((row[ds], i))
        # Sort descending by accuracy to get rank
        vals.sort(key=lambda x: x[0], reverse=True)
        for rank, (val, idx) in enumerate(vals, start=1):
            ranking_rows[idx]['ranks'][ds] = rank
    
    # 2. Calculate Mean Rank across all datasets
    for row in ranking_rows:
        ranks = [r for r in row['ranks'].values()]
        row['GLOBAL_MEAN'] = sum(ranks) / len(ranks) if ranks else None
            
    return ranking_rows
def build_gain_pivot(abs_rows_main, abs_rows_gcn, datasets):
    """Compute (main_mean - gcn_mean) for every cell."""
    gain_rows = []
    for row_m, row_g in zip(abs_rows_main, abs_rows_gcn):
        grow = {'noise_type': row_m['noise_type'], 'noise_rate': row_m['noise_rate']}
        for ds in datasets:
            vm = row_m.get(ds)
            vg = row_g.get(ds)
            if vm is not None and vg is not None:
                mean_m = vm[0]
                mean_g = vg[0]
                grow[ds] = (mean_m - mean_g, None)
            else:
                grow[ds] = None
        gain_rows.append(grow)
    return gain_rows
# ---------------------------------------------------------------------------
# Terminal printing
# ---------------------------------------------------------------------------
def row_label(noise_type, noise_rate):
    nt = NOISE_TYPE_LABELS.get(noise_type, noise_type.capitalize())
    if noise_type == 'MEAN':
        return 'Média'
    if noise_type == 'clean':
        return f'{nt}'
    return f'{nt} {noise_rate:.1f}'
def fmt_abs(cell):
    if cell is None:
        return '  ---  '
    mean, std = cell
    if std is None:
        return f'{mean:6.2f}'
    return f'{mean:6.2f}±{std:.2f}'
def fmt_gain(cell):
    if cell is None:
        return '  ---  '
    gain, _ = cell
    sign = '+' if gain >= 0 else ''
    return f'{sign}{gain:6.2f}'
def print_table(rows, datasets, title, fmt_fn):
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets]
    col_w = max(12, max(len(l) for l in ds_labels) + 2)
    row_w = 15
    print()
    print('=' * (row_w + col_w * len(datasets) + 4))
    print(f'  {title}')
    print('=' * (row_w + col_w * len(datasets) + 4))
    header = f'{"Ruído":<{row_w}}' + ''.join(f'{l:^{col_w}}' for l in ds_labels)
    print(header)
    print('-' * (row_w + col_w * len(datasets) + 4))
    for i, row in enumerate(rows):
        label = row_label(row['noise_type'], row['noise_rate'])
        if row['noise_type'] == 'MEAN' and i > 0:
            print('-' * (row_w + col_w * len(datasets) + 4))
        cells = ''.join(f'{fmt_fn(row.get(ds)):^{col_w}}' for ds in datasets)
        print(f'{label:<{row_w}}{cells}')
    print('=' * (row_w + col_w * len(datasets) + 4))

def print_summary_table(rows, datasets, title):
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets] + ['Média']
    datasets_ext = datasets + ['GLOBAL_MEAN']
    col_w = max(10, max(len(l) for l in ds_labels) + 1)
    row_w = 12
    print(f"\n{'='*80}\n  {title}\n{'='*80}")
    header = f'{"Método":<{row_w}}' + ''.join(f'{l:>{col_w}}' for l in ds_labels)
    print(header)
    print('-' * len(header))
    
    for row in rows:
        m_label = METHOD_LABELS.get(row['method'], row['method'])
        cells = ""
        for ds in datasets_ext:
            val = row.get(ds)
            if val is None:
                cells += f'{"---":>{col_w}}'
            else:
                col_vals = [r[ds] for r in rows if r.get(ds) is not None]
                is_best = abs(val - max(col_vals)) < 1e-5
                cell_str = f"{val:6.2f}"
                if is_best: cell_str = f"*{cell_str}*"
                cells += f'{cell_str:>{col_w}}'
        print(f'{m_label:<{row_w}}{cells}')
    print('=' * len(header))

def print_ranking_table(rows, datasets, title):
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets] + ['Rank Médio']
    datasets_ext = datasets + ['GLOBAL_MEAN']
    col_w = max(11, max(len(l) for l in ds_labels) + 1)
    row_w = 12
    print(f"\n{'='*80}\n  {title}\n{'='*80}")
    header = f'{"Método":<{row_w}}' + ''.join(f'{l:>{col_w}}' for l in ds_labels)
    print(header)
    print('-' * len(header))
    
    for row in rows:
        m_label = METHOD_LABELS.get(row['method'], row['method'])
        cells = ""
        for ds in datasets_ext:
            val = row['ranks'].get(ds) if ds != 'GLOBAL_MEAN' else row.get('GLOBAL_MEAN')
            if val is None:
                cells += f'{"---":>{col_w}}'
            elif ds == 'GLOBAL_MEAN':
                # Bold check (lowest mean rank)
                all_means = [r.get('GLOBAL_MEAN') for r in rows if r.get('GLOBAL_MEAN') is not None]
                is_best = abs(val - min(all_means)) < 1e-5
                s = f"{val:.2f}"
                if is_best: s = f"*{s}*"
                cells += f'{s:>{col_w}}'
            else:
                cells += f'{int(val):>{col_w}}'
        print(f'{m_label:<{row_w}}{cells}')
    print('=' * len(header))
# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def _excel_color(val, is_gain=False):
    """Return openpyxl hex fill colour based on value magnitude."""
    if val is None:
        return None
    if is_gain:
        if val > 2.0:   return 'C6EFCE'   # green
        if val > 0.5:   return 'FFEB9C'   # yellow
        if val > -0.5:  return None
        return 'FFC7CE'                    # red
    return None
def write_excel_sheet(wb, sheet_name, rows, datasets, is_gain=False, title=''):
    ws = wb.create_sheet(title=sheet_name)
    # ── Styles ──────────────────────────────────────────────────────────────
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    mean_fill = PatternFill("solid", fgColor="D9E1F2")
    bold_font = Font(bold=True, size=10)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets]
    # ── Title row ───────────────────────────────────────────────────────────
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(datasets))
    # ── Header ──────────────────────────────────────────────────────────────
    header_row = ['Ruído'] + ds_labels
    ws.append(header_row)
    for col_idx, val in enumerate(header_row, start=1):
        cell = ws.cell(2, col_idx, val)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = center
        cell.border = border
    # ── Data rows ───────────────────────────────────────────────────────────
    for row in rows:
        is_mean_row = row['noise_type'] == 'MEAN'
        label = row_label(row['noise_type'], row['noise_rate'])
        excel_row = [label]
        for ds in datasets:
            cell_val = row.get(ds)
            if cell_val is None:
                excel_row.append(None)
            else:
                mean, std = cell_val
                if is_gain:
                    excel_row.append(round(mean, 2))
                else:
                    if std is not None:
                        excel_row.append(round(mean, 2))  # mean only for compactness
                    else:
                        excel_row.append(round(mean, 2))
        ws.append(excel_row)
        r = ws.max_row
        for col_idx in range(1, len(excel_row) + 1):
            c = ws.cell(r, col_idx)
            c.border = border
            c.alignment = center if col_idx > 1 else left_align
            if is_mean_row:
                c.fill = mean_fill
                c.font = bold_font
            if col_idx > 1:
                raw_val = row.get(datasets[col_idx - 2])
                if raw_val is not None:
                    color = _excel_color(raw_val[0], is_gain=is_gain)
                    if color:
                        c.fill = PatternFill("solid", fgColor=color)
                        if is_gain and raw_val[0] > 2.0:
                            c.font = Font(bold=True, size=10)
        # Add ± std in comment or second sub-column? Keep simple: just mean.
    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 18
    for i in range(len(datasets)):
        col_letter = get_column_letter(i + 2)
        ws.column_dimensions[col_letter].width = 12
    # ── Freeze panes ────────────────────────────────────────────────────────
    ws.freeze_panes = 'B3'

def write_summary_sheet(wb, rows, datasets, title):
    ws = wb.create_sheet(title='Resumo Médio')
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    best_fill = PatternFill("solid", fgColor="C6EFCE") # light green for bold
    bold_font = Font(bold=True, size=10)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets] + ['Média']
    datasets_ext = datasets + ['GLOBAL_MEAN']
    
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.append(['Método'] + ds_labels)
    for col_idx in range(1, len(ds_labels) + 2):
        c = ws.cell(2, col_idx)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = border
        
    for row in rows:
        m_label = METHOD_LABELS.get(row['method'], row['method'])
        excel_row = [m_label]
        for ds in datasets_ext:
            excel_row.append(row.get(ds))
        ws.append(excel_row)
        r = ws.max_row
        for col_idx in range(1, len(excel_row) + 1):
            c = ws.cell(r, col_idx)
            c.border = border
            c.alignment = center if col_idx > 1 else left_align
            if col_idx > 1:
                val = row.get(datasets_ext[col_idx - 2])
                if val is not None:
                    col_vals = [r[datasets_ext[col_idx - 2]] for r in rows if r.get(datasets_ext[col_idx - 2]) is not None]
                    if abs(val - max(col_vals)) < 1e-5:
                        c.font = bold_font
                        c.fill = best_fill
    ws.column_dimensions['A'].width = 15

def write_ranking_sheet(wb, rows, datasets, title):
    ws = wb.create_sheet(title='Ranking')
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    rank1_fill = PatternFill("solid", fgColor="FFD700") # Gold for Rank 1
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets] + ['Rank Médio']
    datasets_ext = datasets + ['GLOBAL_MEAN']
    
    ws.append([title])
    ws.cell(1, 1).font = Font(bold=True, size=12)
    ws.append(['Método'] + ds_labels)
    for col_idx in range(1, len(ds_labels) + 2):
        c = ws.cell(2, col_idx)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = border
        
    for row in rows:
        m_label = METHOD_LABELS.get(row['method'], row['method'])
        excel_row = [m_label]
        for ds in datasets_ext:
            if ds == 'GLOBAL_MEAN':
                val = row.get('GLOBAL_MEAN')
                excel_row.append(val)
            else:
                excel_row.append(row['ranks'].get(ds))
        ws.append(excel_row)
        r = ws.max_row
        for col_idx in range(1, len(excel_row) + 1):
            c = ws.cell(r, col_idx)
            c.border = border
            c.alignment = center if col_idx > 1 else left_align
            # Highlight dataset Rank 1
            if 1 < col_idx <= len(datasets) + 1:
                if excel_row[col_idx-1] == 1:
                    c.fill = rank1_fill
                    c.font = Font(bold=True)
            # Highlight BEST Mean Rank (lowest)
            if ds_labels[col_idx-2] == 'Rank Médio' if col_idx > 1 else False:
                val = excel_row[col_idx-1]
                all_means = [r.get('GLOBAL_MEAN') for r in rows if r.get('GLOBAL_MEAN') is not None]
                if val is not None and abs(val - min(all_means)) < 1e-5:
                    c.font = Font(bold=True)
                    c.fill = rank1_fill # Highlight best mean rank too
    ws.column_dimensions['A'].width = 15
def write_hyperparam_sheet(wb, hyperparams: dict):
    """Write a sheet with per-dataset hyperparameter summary (clean scenario)."""
    ws = wb.create_sheet(title='HP por Dataset')
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left', vertical='center')
    datasets_present = [d for d in DATASET_ORDER if d in hyperparams]
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets_present]
    params = list(HYPERPARAM_DISPLAY.keys())
    # Title
    ws.append(['HPO (cenário clean por dataset)'])
    ws.cell(1, 1).font = Font(bold=True, size=12)
    # Header
    header = ['Parâmetro'] + ds_labels
    ws.append(header)
    for col_idx, val in enumerate(header, start=1):
        c = ws.cell(2, col_idx, val)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = center
        c.border = border
    # Rows
    for param in params:
        label = HYPERPARAM_DISPLAY[param]
        row = [label]
        for ds in datasets_present:
            val = hyperparams.get(ds, {}).get(param)
            row.append(_fmt_hpo_val(param, val))
        ws.append(row)
        r = ws.max_row
        for col_idx in range(1, len(row) + 1):
            c = ws.cell(r, col_idx)
            c.border = border
            c.alignment = center if col_idx > 1 else left_align
    ws.column_dimensions['A'].width = 16
    for i in range(len(datasets_present)):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14

def write_hpo_scenarios_sheet(wb, hpo_by_ds: dict, datasets: list,
                              noise_types: list, noise_rates: list):
    """
    Write a per-scenario HPO sheet.
    Rows: (noise_type, noise_rate) — same structure as result tables.
    Columns: grouped by dataset, one sub-column per hyperparameter.
    Layout: Ruído | Rate || [Dataset1: param1 param2 ...] || [Dataset2: ...] || ...
    """
    ws = wb.create_sheet(title='HP por Cenário')
    hdr_fill  = PatternFill("solid", fgColor="1F3864")
    hdr2_fill = PatternFill("solid", fgColor="2F5597")
    hdr_font  = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal='center', vertical='center')
    left_align = Alignment(horizontal='left',   vertical='center')

    params       = list(HYPERPARAM_DISPLAY.keys())
    param_labels = list(HYPERPARAM_DISPLAY.values())
    ds_present   = [d for d in datasets if d in hpo_by_ds]
    n_params     = len(params)

    # ── Row 1: dataset group headers ────────────────────────────────────────
    ws.cell(1, 1, 'Ruído').fill = hdr_fill
    ws.cell(1, 1).font = hdr_font
    ws.cell(1, 1).alignment = center
    ws.cell(1, 1).border = border
    ws.cell(1, 2, 'Rate').fill = hdr_fill
    ws.cell(1, 2).font = hdr_font
    ws.cell(1, 2).alignment = center
    ws.cell(1, 2).border = border
    col = 3
    for ds in ds_present:
        ds_label = DATASET_LABELS.get(ds, ds)
        ws.cell(1, col, ds_label).fill = hdr_fill
        ws.cell(1, col).font = hdr_font
        ws.cell(1, col).alignment = center
        ws.cell(1, col).border = border
        if n_params > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1,   end_column=col + n_params - 1)
        col += n_params

    # ── Row 2: param sub-headers ─────────────────────────────────────────────
    ws.cell(2, 1).border = border
    ws.cell(2, 2).border = border
    col = 3
    for ds in ds_present:
        for pl in param_labels:
            c = ws.cell(2, col, pl)
            c.fill = hdr2_fill
            c.font = hdr_font
            c.alignment = center
            c.border = border
            col += 1

    # ── Data rows ────────────────────────────────────────────────────────────
    row_idx = 3
    prev_nt = None
    for nt in noise_types:
        rates = [0.0] if nt == 'clean' else [nr for nr in noise_rates if nr > 0.0]
        for nr in rates:
            nt_label = NOISE_TYPE_LABELS.get(nt, nt.capitalize())
            nr_label = 'Clean' if nt == 'clean' else f'{nr:.1f}'
            # Group separator
            if nt != prev_nt and prev_nt is not None:
                # light separator row for visual grouping (just bold borders)
                pass
            c1 = ws.cell(row_idx, 1, nt_label if nt != prev_nt else '')
            c1.border = border
            c1.alignment = left_align
            c2 = ws.cell(row_idx, 2, nr_label)
            c2.border = border
            c2.alignment = center
            col = 3
            for ds in ds_present:
                scenario = hpo_by_ds.get(ds, {}).get((nt, nr))
                for param in params:
                    val = scenario.get(param) if scenario else None
                    cell_str = _fmt_hpo_val(param, val)
                    c = ws.cell(row_idx, col, cell_str)
                    c.border = border
                    c.alignment = center
                    col += 1
            prev_nt = nt
            row_idx += 1

    # ── Column widths ────────────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 7
    col_letter_idx = 3
    for ds in ds_present:
        for param in params:
            ltr = get_column_letter(col_letter_idx)
            # knn_mode and k can be narrower
            ws.column_dimensions[ltr].width = 7 if param in ('knn_mode', 'k', 'n_layer', 'n_hidden') else 8
            col_letter_idx += 1
    ws.freeze_panes = 'C3'
# ---------------------------------------------------------------------------
# LaTeX export
# ---------------------------------------------------------------------------
def latex_cell_abs(cell, bold=False):
    if cell is None:
        return '---'
    mean, std = cell
    if std is None:
        s = f'{mean:.2f}'
    else:
        s = f'{mean:.2f}{{\\scriptscriptstyle\\pm}}{std:.2f}'
    s = f'${s}$'
    return f'\\textbf{{{s}}}' if bold else s
def latex_cell_gain(cell, bold=False):
    if cell is None:
        return '---'
    gain, _ = cell
    sign = '+' if gain >= 0 else ''
    s = f'{sign}{gain:.2f}'
    return f'\\textbf{{{s}}}' if bold else s
def write_latex(path: Path, rows_abs, rows_gain, rows_abs_gcn,
                datasets, method_label, gcn_label='GCN'):
    """Write a .tex file with two tables: absolute and gain."""
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets]
    def make_table(rows, fmt_fn, caption, label, is_gain):
        n_ds = len(datasets)
        col_spec = 'll' + 'c' * n_ds
        ds_header = ' & '.join(f'\\textbf{{{l}}}' for l in ds_labels)
        lines = []
        lines.append(r'\begin{table*}[htbp]')
        lines.append(r'\centering')
        lines.append(r'\small')
        lines.append(f'\\caption{{{caption}}}')
        lines.append(f'\\label{{{label}}}')
        lines.append(r'\resizebox{\textwidth}{!}{%')
        lines.append(f'\\begin{{tabular}}{{{col_spec}}}')
        lines.append(r'\toprule')
        lines.append(f'\\textbf{{Noise}} & \\textbf{{Rate}} & {ds_header} \\\\')
        lines.append(r'\midrule')
        prev_nt = None
        for row in rows:
            nt = row['noise_type']
            nr = row.get('noise_rate')
            is_mean = nt == 'MEAN'
            if is_mean:
                lines.append(r'\midrule')
            if is_mean:
                nt_cell = '\\textbf{Média}'
                nr_cell = ''
            elif nt != prev_nt:
                nt_cell = NOISE_TYPE_LABELS.get(nt, nt.capitalize())
                nr_cell = 'Clean' if nt == 'clean' else f'{nr:.1f}'
            else:
                nt_cell = ''
                nr_cell = f'{nr:.1f}'
            # Determine best column for bold
            raw_vals = [row.get(ds) for ds in datasets]
            valid_vals = [v[0] for v in raw_vals if v is not None]
            if is_gain:
                best = max(valid_vals) if valid_vals else None
            else:
                best = max(valid_vals) if valid_vals else None
            cells = []
            for ds in datasets:
                cell = row.get(ds)
                is_best = (
                    cell is not None and best is not None and
                    abs(cell[0] - best) < 1e-4 and not is_mean
                )
                cells.append(fmt_fn(cell, bold=is_best))
            lines.append(
                f'{nt_cell} & {nr_cell} & ' + ' & '.join(cells) + r' \\'
            )
            prev_nt = nt
        lines.append(r'\bottomrule')
        lines.append(r'\end{tabular}')
        lines.append(r'}')
        lines.append(r'\end{table*}')
        lines.append('')
        return '\n'.join(lines)
    abs_table = make_table(
        rows_abs,
        latex_cell_abs,
        caption=f'{method_label} — Acurácia absoluta (\\%)',
        label='tab:abs_results',
        is_gain=False,
    )
    gain_table = make_table(
        rows_gain,
        latex_cell_gain,
        caption=f'{method_label} — Ganho relativo ao {gcn_label} (pontos percentuais)',
        label='tab:gain_results',
        is_gain=True,
    )
    preamble = (
        '% Generated by results/generate_tables.py\n'
        '% Required: \\usepackage{booktabs, multirow, graphicx}\n\n'
    )
    with open(path, 'w', encoding='utf-8') as f:
        f.write(preamble + abs_table + '\n\n' + gain_table)

def write_latex_summary(path: Path, summary_rows, ranking_rows, datasets):
    """Append summary and ranking tables to the .tex file."""
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets] + ['Média']
    datasets_ext = datasets + ['GLOBAL_MEAN']
    n_cols = len(datasets_ext)
    
    def make_summary_tex():
        lines = [r'\begin{table*}[htbp]', r'\centering', r'\small', 
                 r'\caption{Resumo: Acurácia Média Global por Método (Clean + Noisy)}', 
                 r'\label{tab:summary_mean}', r'\begin{tabular}{l' + 'c'*n_cols + '}', r'\toprule',
                 r'\textbf{Método} & ' + ' & '.join(f'\\textbf{{{l}}}' for l in ds_labels) + r' \\', r'\midrule']
        for row in summary_rows:
            m_label = METHOD_LABELS.get(row['method'], row['method']).replace('_', '-')
            cells = []
            for ds in datasets_ext:
                val = row.get(ds)
                if val is None: cells.append('---')
                else:
                    col_vals = [r[ds] for r in summary_rows if r.get(ds) is not None]
                    is_best = abs(val - max(col_vals)) < 1e-5
                    s = f"{val:.2f}"
                    cells.append(f'\\textbf{{{s}}}' if is_best else s)
            lines.append(f'{m_label} & ' + ' & '.join(cells) + r' \\')
        lines.extend([r'\bottomrule', r'\end{tabular}', r'\end{table*}', ''])
        return '\n'.join(lines)

    def make_ranking_tex():
        ds_labels_tex = [DATASET_LABELS.get(d, d) for d in datasets] + ['Rank Médio']
        lines = [r'\begin{table*}[htbp]', r'\centering', r'\small', 
                 r'\caption{Ranking Médio por Dataset (Posição na Tabela 1)}', 
                 r'\label{tab:summary_ranking}', r'\begin{tabular}{l' + 'c'*n_cols + '}', r'\toprule',
                 r'\textbf{Método} & ' + ' & '.join(f'\\textbf{{{l}}}' for l in ds_labels_tex) + r' \\', r'\midrule']
        for row in ranking_rows:
            m_label = METHOD_LABELS.get(row['method'], row['method']).replace('_', '-')
            cells = []
            for ds in datasets_ext:
                if ds == 'GLOBAL_MEAN':
                    val = row.get('GLOBAL_MEAN')
                    if val is None: cells.append('---')
                    else:
                        all_means = [r.get('GLOBAL_MEAN') for r in ranking_rows if r.get('GLOBAL_MEAN') is not None]
                        is_best = abs(val - min(all_means)) < 1e-5
                        s = f"{val:.2f}"
                        cells.append(f'\\textbf{{{s}}}' if is_best else s)
                else:
                    rank = row['ranks'].get(ds)
                    cells.append(str(rank) if rank is not None else '---')
            lines.append(f'{m_label} & ' + ' & '.join(cells) + r' \\')
        lines.extend([r'\bottomrule', r'\end{tabular}', r'\end{table*}', ''])
        return '\n'.join(lines)

    with open(path, 'a', encoding='utf-8') as f:
        f.write('\n\n' + make_summary_tex() + '\n\n' + make_ranking_tex())
def write_latex_hyperparams(path: Path, hyperparams: dict):
    """Write a .tex file with the hyperparameter table."""
    datasets_present = [d for d in DATASET_ORDER if d in hyperparams]
    ds_labels = [DATASET_LABELS.get(d, d) for d in datasets_present]
    params = list(HYPERPARAM_DISPLAY.keys())
    param_labels = [HYPERPARAM_DISPLAY[p] for p in params]
    col_spec = 'l' + 'c' * len(datasets_present)
    ds_header = ' & '.join(f'\\textbf{{{l}}}' for l in ds_labels)
    lines = []
    lines.append(r'\begin{table}[htbp]')
    lines.append(r'\centering')
    lines.append(r'\small')
    lines.append(r'\caption{LN-PCC Hyperparâmetros ajustados por dataset}')
    lines.append(r'\label{tab:hyperparams}')
    lines.append(r'\resizebox{\textwidth}{!}{%')
    lines.append(f'\\begin{{tabular}}{{{col_spec}}}')
    lines.append(r'\toprule')
    lines.append(f'\\textbf{{Parâmetro}} & {ds_header} \\\\')
    lines.append(r'\midrule')
    for param, label in zip(params, param_labels):
        cells = []
        for ds in datasets_present:
            val = hyperparams.get(ds, {}).get(param)
            if val is None:
                cells.append('---')
            elif isinstance(val, float):
                cells.append(f'{val:.4g}')
            else:
                cells.append(str(val))
        lines.append(f'{label} & ' + ' & '.join(cells) + r' \\')
    lines.append(r'\bottomrule')
    lines.append(r'\end{tabular}')
    lines.append(r'}')
    lines.append(r'\end{table}')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('% Generated by results/generate_tables.py\n')
        f.write('% Required: \\usepackage{booktabs, graphicx}\n\n')
        f.write('\n'.join(lines))
# ---------------------------------------------------------------------------
# Load hyperparameters from HPO database (hpo_db.json)
# ---------------------------------------------------------------------------
def _fmt_hpo_val(param: str, val) -> str:
    """Format a single HPO parameter value for display."""
    if val is None:
        return '—'
    if param in HYPERPARAM_INT:
        return str(int(round(val))) if isinstance(val, float) else str(val)
    if param in HYPERPARAM_FLOAT4:
        try:
            return f'{float(val):.4g}'
        except (TypeError, ValueError):
            return str(val)
    return str(val)

def load_hpo_db(hpo_db_path: Path) -> dict:
    """
    Load hpo_db.json.
    Returns
    -------
    dict: {
      dataset: {
        (noise_type, noise_rate): {param: value, ...},
        ...
      },
      ...
    }
    """
    if not hpo_db_path.exists():
        print(f'[AVISO] HPO DB não encontrada: {hpo_db_path}')
        return {}
    import json
    with open(hpo_db_path, encoding='utf-8') as f:
        raw = json.load(f)
    result = {}
    for key, params in raw.items():
        # key format: "dataset_noisetype_rate"  e.g. "cora_uniform_0.3"
        # Split from right to get rate, then from right again for noise_type
        parts = key.rsplit('_', 1)
        if len(parts) != 2:
            continue
        prefix, rate_str = parts
        try:
            rate = float(rate_str)
        except ValueError:
            continue
        # prefix is like "cora_uniform" or "amazon-ratings_pair"
        # noise types we know:
        noise_types_known = ['clean', 'uniform', 'pair', 'random', 'instance']
        noise_type = None
        dataset = None
        for nt in noise_types_known:
            if prefix.endswith('_' + nt):
                noise_type = nt
                dataset = prefix[: -(len(nt) + 1)]
                break
        if noise_type is None or dataset is None:
            continue
        if dataset not in result:
            result[dataset] = {}
        result[dataset][(noise_type, rate)] = {k: v for k, v in params.items()
                                                if k in HYPERPARAM_DISPLAY}
    return result

def hpo_db_to_dataset_summary(hpo_by_ds: dict) -> dict:
    """
    Collapse hpo_by_ds to a per-dataset dict using the 'clean' scenario
    (or first available scenario as fallback).
    Returns dict: {dataset: {param: value}}
    """
    result = {}
    for ds, scenarios in hpo_by_ds.items():
        # Prefer clean scenario as representative
        chosen = scenarios.get(('clean', 0.0))
        if chosen is None and scenarios:
            chosen = next(iter(scenarios.values()))
        result[ds] = chosen or {}
    return result

# Keep load_hyperparams as a thin wrapper for backwards compat
def load_hyperparams(config_dir: Path) -> dict:
    """Legacy: load YAML-based hyperparams. Kept for fallback."""
    return {}  # No longer used as primary source
# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def parse_args():
    p = argparse.ArgumentParser(
        description='Gera tabelas de resultados LN-PCC (terminal, Excel, LaTeX)'
    )
    p.add_argument(
        '--csv', type=str, default=None,
        help='Caminho para o CSV de resultados. '
             'Se omitido, usa o lnpcc_results_*.csv mais recente em log/'
    )
    p.add_argument(
        '--methods', nargs='+', default=['lnpcc', 'gcn'],
        help='Métodos a incluir nas tabelas (default: lnpcc gcn)'
    )
    p.add_argument(
        '--main_method', type=str, default='lnpcc',
        help='Método principal para ganho relativo (default: lnpcc)'
    )
    p.add_argument(
        '--baseline', type=str, default='gcn',
        help='Método baseline para cálculo de ganho (default: gcn)'
    )
    p.add_argument(
        '--datasets', nargs='+', default=None,
        help='Datasets a incluir (default: todos encontrados no CSV)'
    )
    p.add_argument(
        '--noise_types', nargs='+', default=None,
        choices=list(NOISE_TYPE_LABELS.keys()),
        help='Tipos de ruído a incluir (default: todos)'
    )
    p.add_argument(
        '--noise_rates', nargs='+', type=float, default=None,
        help='Taxas de ruído a incluir (default: todas)'
    )
    p.add_argument(
        '--outdir', type=str, default=None,
        help='Diretório de saída (default: mesmo diretório deste script)'
    )
    p.add_argument(
        '--no_excel', action='store_true',
        help='Não gerar o arquivo Excel'
    )
    p.add_argument(
        '--no_latex', action='store_true',
        help='Não gerar o arquivo LaTeX'
    )
    p.add_argument(
        '--no_hyperparams', action='store_true',
        help='Não gerar a tabela de hiperparâmetros'
    )
    return p.parse_args()
# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    args = parse_args()
    # ── Output directory ────────────────────────────────────────────────────
    outdir = Path(args.outdir) if args.outdir else SCRIPT_DIR / "results"
    outdir.mkdir(parents=True, exist_ok=True)
    # ── CSV ─────────────────────────────────────────────────────────────────
    if args.csv:
        csv_path = Path(args.csv)
    else:
        csv_path = find_latest_csv(LOG_DIR)
        print(f'[INFO] CSV encontrado automaticamente: {csv_path}')
    print(f'[INFO] Carregando dados de: {csv_path}')
    data = load_csv(csv_path)
    
    # ── External Baselines ──────────────────────────────────────────────────
    baseline_path = PROJECT_DIR / "results" / "NoisyGL.xlsx"
    excel_data = load_external_baseline_xlsx(baseline_path)
    if excel_data:
        print(f'[INFO] Carregando baselines de: {baseline_path.name} ({len(excel_data)} registros)')
        data.update(excel_data)
    
    if not data:
        print('ERRO: Nenhum dado carregado. Verifique os CSVs.', file=sys.stderr)
        sys.exit(1)

    # ── Global Parameters ───────────────────────────────────────────────────
    all_methods_in_data = sorted(list(set(k[0] for k in data.keys())))
    all_datasets = sorted({k[1] for k in data})
    datasets = args.datasets or all_datasets
    # Canonical order
    datasets = [d for d in DATASET_ORDER if d in datasets]
    
    # Filter 'clean' if not requested
    noise_types = args.noise_types or NOISE_TYPE_ORDER
    noise_rates = args.noise_rates or [0.0, 0.1, 0.2, 0.3, 0.4, 0.5]
    
    main_method = args.main_method
    baseline    = args.baseline

    # ── Build Standard Tables ────────────────────────────────────────────────
    rows_main = build_pivot(data, main_method, datasets, noise_types, noise_rates)
    rows_gcn  = build_pivot(data, baseline, datasets, noise_types, noise_rates)
    rows_gain = build_gain_pivot(rows_main, rows_gcn, datasets)
    main_label = METHOD_LABELS.get(main_method, main_method.upper())
    gcn_label  = METHOD_LABELS.get(baseline, baseline.upper())

    # ── Build Summary Tables ─────────────────────────────────────────────────
    summary_rows = build_summary_table(data, all_methods_in_data, datasets)
    ranking_rows = build_ranking_table(summary_rows, datasets)

    # ── Terminal Output ─────────────────────────────────────────────────────
    print_table(rows_main, datasets, f'{main_label} — Acurácia Absoluta (%)', fmt_abs)
    print_table(rows_gain, datasets, f'Ganho de {main_label} relativo ao {gcn_label} (pp)', fmt_gain)
    
    print_summary_table(summary_rows, datasets, "Resumo: Acurácia Média Global por Método (Clean + Noisy)")
    print_ranking_table(ranking_rows, datasets, "Resumo: Ranking Médio por Dataset")

    # ── Excel Export ────────────────────────────────────────────────────────
    if not args.no_excel:
        if not HAS_OPENPYXL:
            print('[AVISO] openpyxl não instalado — Excel desativado.')
        else:
            out_xlsx = outdir / 'results.xlsx'
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            write_excel_sheet(wb, f'{main_label} Absoluto', rows_main, datasets, False, f'{main_label} Acurácia Absoluta (%)')
            write_excel_sheet(wb, f'Ganho vs {gcn_label}', rows_gain, datasets, True, f'Ganho de {main_label} vs {gcn_label} (pp)')
            
            write_summary_sheet(wb, summary_rows, datasets, "Acurácia Média (Todos os cenários)")
            write_ranking_sheet(wb, ranking_rows, datasets, "Ranking por Dataset")

            hpo_db = load_hpo_db(HPO_DB_PATH)
            hpo_summary = hpo_db_to_dataset_summary(hpo_db)
            if hpo_summary:
                write_hyperparam_sheet(wb, hpo_summary)
                write_hpo_scenarios_sheet(wb, hpo_db, datasets, noise_types, noise_rates)
            
            wb.save(out_xlsx)
            print(f'\n[OK] Excel salvo em: {out_xlsx}')

    # ── LaTeX Export ─────────────────────────────────────────────────────────
    if not args.no_latex:
        out_tex = outdir / 'results.tex'
        write_latex(out_tex, rows_main, rows_gain, rows_gcn, datasets, main_label, gcn_label)
        write_latex_summary(out_tex, summary_rows, ranking_rows, datasets)
        print(f'[OK] LaTeX salvo em: {out_tex}')
        
        if hpo_summary:
            hp_tex_path = outdir / 'hyperparams.tex'
            write_latex_hyperparams(hp_tex_path, hpo_summary)

    print("\nProcessamento concluído com sucesso!")

if __name__ == "__main__":
    main()
